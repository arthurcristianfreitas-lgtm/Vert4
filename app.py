# app.py — Flask completo com tracking + admin analytics
from flask import Flask, render_template, request, jsonify, send_file
from flask_sqlalchemy import SQLAlchemy
from datetime import datetime, timedelta
import anthropic, json, os, io, uuid
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

load_dotenv()

app = Flask(__name__)
app.config["DEBUG"]         = os.getenv("FLASK_DEBUG","true").lower()=="true"
app.config["SQLALCHEMY_DATABASE_URI"] = "sqlite:///vert4_analytics.db"
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False

from database import db, Session, Event
db.init_app(app)

with app.app_context():
    db.create_all()

client = anthropic.Anthropic(api_key=os.getenv("ANTHROPIC_API_KEY"))
ADMIN_PASSWORD = os.getenv("ADMIN_PASSWORD", "VERT4@2025")

def vite_asset(filename):
    if app.config["DEBUG"]:
        return f"http://localhost:5173/{filename}"
    try:
        with open(os.path.join(app.static_folder,".vite","manifest.json")) as f:
            manifest = json.load(f)
        return "/static/" + manifest[filename]["file"]
    except:
        return f"/static/{filename}"

app.jinja_env.globals["vite_asset"] = vite_asset

# ── FRONTEND ──────────────────────────────────────────────────────────────────
@app.route("/")
def index():
    return render_template("index.html")

# ── API: ANÁLISE ──────────────────────────────────────────────────────────────
@app.route("/api/health", methods=["GET"])
def health():
    return jsonify({"status":"ok","api_key_configurada":bool(os.getenv("ANTHROPIC_API_KEY"))})

@app.route("/api/analyze", methods=["POST"])
def analyze():
    data   = request.get_json(silent=True) or {}
    prompt = data.get("prompt","").strip()
    if not prompt:
        return jsonify({"error":"Prompt não informado"}), 400
    if not os.getenv("ANTHROPIC_API_KEY"):
        return jsonify({"error":"ANTHROPIC_API_KEY não configurada"}), 500
    try:
        msg = client.messages.create(
            model="claude-sonnet-4-20250514", max_tokens=1000,
            messages=[{"role":"user","content":prompt}]
        )
        return jsonify({"content":[{"type":"text","text":msg.content[0].text}]})
    except anthropic.AuthenticationError:
        return jsonify({"error":"Chave inválida"}), 401
    except Exception as e:
        return jsonify({"error":str(e)}), 500

# ── API: TRACKING ─────────────────────────────────────────────────────────────
@app.route("/api/track/consent", methods=["POST"])
def track_consent():
    """Cria sessão quando usuário aceita os cookies."""
    data       = request.get_json(silent=True) or {}
    session_id = data.get("session_id") or str(uuid.uuid4())
    consented  = data.get("consented", False)

    sess = Session.query.get(session_id)
    if not sess:
        sess = Session(
            id         = session_id,
            ip         = request.remote_addr,
            user_agent = request.user_agent.string[:256],
            consented  = consented,
        )
        db.session.add(sess)
    else:
        sess.consented  = consented
        sess.last_seen  = datetime.utcnow()
    db.session.commit()
    return jsonify({"session_id": session_id, "ok": True})

@app.route("/api/track/event", methods=["POST"])
def track_event():
    """Registra um evento de comportamento."""
    data       = request.get_json(silent=True) or {}
    session_id = data.get("session_id")
    event_type = data.get("event_type")   # page_view | tab_view | cta_click | form_start | form_complete | time_update
    payload    = data.get("payload","")

    if not session_id or not event_type:
        return jsonify({"error":"session_id e event_type obrigatórios"}), 400

    sess = Session.query.get(session_id)
    if not sess or not sess.consented:
        return jsonify({"error":"Sessão não encontrada ou sem consentimento"}), 403

    # Atualiza flags especiais
    if event_type == "form_complete":
        sess.completed_diag = True
    if event_type in ("tab_view","cta_click") and "fase2" in str(payload).lower():
        sess.purchase_intent = True
    if event_type == "cta_click" and "whatsapp" in str(payload).lower():
        sess.purchase_intent = True
    if event_type == "time_update":
        try:
            sess.duration_sec = int(json.loads(payload).get("seconds", 0))
        except:
            pass
    sess.last_seen = datetime.utcnow()

    ev = Event(session_id=session_id, event_type=event_type,
               payload=json.dumps(payload) if isinstance(payload,dict) else str(payload))
    db.session.add(ev)
    db.session.commit()
    return jsonify({"ok": True})

# ── API: ADMIN ────────────────────────────────────────────────────────────────
@app.route("/api/admin/login", methods=["POST"])
def admin_login():
    data = request.get_json(silent=True) or {}
    if data.get("password") == ADMIN_PASSWORD:
        return jsonify({"ok":True})
    return jsonify({"error":"Senha incorreta"}), 401

@app.route("/api/admin/stats", methods=["POST"])
def admin_stats():
    data = request.get_json(silent=True) or {}
    if data.get("password") != ADMIN_PASSWORD:
        return jsonify({"error":"Não autorizado"}), 401

    sessions = Session.query.filter_by(consented=True).all()
    total    = len(sessions)
    intents  = sum(1 for s in sessions if s.purchase_intent)
    completed= sum(1 for s in sessions if s.completed_diag)
    avg_dur  = (sum(s.duration_sec for s in sessions) / total) if total else 0

    # Contagem de abas mais vistas
    tab_events = Event.query.filter_by(event_type="tab_view").all()
    tab_count  = {}
    for ev in tab_events:
        try:
            name = json.loads(ev.payload) if ev.payload.startswith('"') else ev.payload
        except:
            name = ev.payload
        tab_count[name] = tab_count.get(name,0) + 1

    # Sessões recentes (últimas 50)
    recent = Session.query.filter_by(consented=True)\
               .order_by(Session.started_at.desc()).limit(50).all()

    sessions_list = [{
        "id":          s.id[:8],
        "ip":          s.ip,
        "started":     s.started_at.strftime("%d/%m/%Y %H:%M"),
        "duration":    s.duration_sec,
        "completed":   s.completed_diag,
        "intent":      s.purchase_intent,
    } for s in recent]

    return jsonify({
        "total_visitors":   total,
        "purchase_intent":  intents,
        "completed_diag":   completed,
        "avg_duration_sec": round(avg_dur),
        "tab_counts":       tab_count,
        "sessions":         sessions_list,
    })

@app.route("/api/admin/export", methods=["POST"])
def admin_export():
    """Exporta dados para Excel."""
    data = request.get_json(silent=True) or {}
    if data.get("password") != ADMIN_PASSWORD:
        return jsonify({"error":"Não autorizado"}), 401

    sessions = Session.query.filter_by(consented=True)\
                 .order_by(Session.started_at.desc()).all()

    wb = Workbook()

    # ── ABA 1: Resumo ──────────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Resumo"

    # Estilos
    hdr_font  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    hdr_fill  = PatternFill("solid", start_color="122920")
    gold_fill = PatternFill("solid", start_color="C8A96E")
    gold_font = Font(name="Arial", bold=True, color="122920", size=11)
    center    = Alignment(horizontal="center", vertical="center")
    thin      = Border(
        left=Side(style="thin",color="2D4A3E"), right=Side(style="thin",color="2D4A3E"),
        top=Side(style="thin",color="2D4A3E"),  bottom=Side(style="thin",color="2D4A3E")
    )

    total     = len(sessions)
    intents   = sum(1 for s in sessions if s.purchase_intent)
    completed = sum(1 for s in sessions if s.completed_diag)
    avg_dur   = int(sum(s.duration_sec for s in sessions)/total) if total else 0

    ws1.merge_cells("A1:D1")
    ws1["A1"] = "VERT4® — Analytics do Concierge Clínico™"
    ws1["A1"].font  = Font(name="Arial", bold=True, color="C8A96E", size=14)
    ws1["A1"].fill  = PatternFill("solid", start_color="0D1F1A")
    ws1["A1"].alignment = center
    ws1.row_dimensions[1].height = 28

    ws1["A3"] = "Indicador";     ws1["B3"] = "Valor"
    for c in ["A3","B3"]:
        ws1[c].font = hdr_font; ws1[c].fill = hdr_fill; ws1[c].alignment = center
    rows = [
        ("Total de visitantes com consentimento", total),
        ("Diagnósticos completos",                completed),
        ("Visitantes com intenção de compra",      intents),
        ("Taxa de intenção (%)",                   f"=B6/B4*100" if total else 0),
        ("Tempo médio no site (segundos)",         avg_dur),
        ("Gerado em",                              datetime.utcnow().strftime("%d/%m/%Y %H:%M UTC")),
    ]
    for i,(k,v) in enumerate(rows, start=4):
        ws1[f"A{i}"] = k; ws1[f"B{i}"] = v
        ws1[f"A{i}"].font = Font(name="Arial", size=10)
        ws1[f"B{i}"].font = Font(name="Arial", bold=True, size=10)
        ws1[f"B{i}"].alignment = center
        for c in [f"A{i}",f"B{i}"]: ws1[c].border = thin
    ws1.column_dimensions["A"].width = 38
    ws1.column_dimensions["B"].width = 22

    # ── ABA 2: Sessões ─────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Sessões")
    headers = ["ID","IP","Iniciou em","Última Ação","Duração (s)","Completou Diagnóstico","Intenção de Compra"]
    for col, h in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = hdr_font; cell.fill = hdr_fill
        cell.alignment = center; cell.border = thin
    for row, s in enumerate(sessions, start=2):
        vals = [
            s.id[:8],
            s.ip,
            s.started_at.strftime("%d/%m/%Y %H:%M"),
            s.last_seen.strftime("%d/%m/%Y %H:%M"),
            s.duration_sec,
            "Sim" if s.completed_diag   else "Não",
            "Sim" if s.purchase_intent  else "Não",
        ]
        for col, val in enumerate(vals, 1):
            c = ws2.cell(row=row, column=col, value=val)
            c.font   = Font(name="Arial", size=10)
            c.border = thin
            c.alignment = center
            if col == 7 and val == "Sim":
                c.fill = PatternFill("solid", start_color="1A8A5A")
                c.font = Font(name="Arial", bold=True, color="FFFFFF")
    for col in range(1,8):
        ws2.column_dimensions[get_column_letter(col)].width = 20

    # ── ABA 3: Eventos ─────────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Eventos")
    ev_headers = ["Sessão ID","Tipo de Evento","Detalhes","Data/Hora"]
    for col, h in enumerate(ev_headers, 1):
        cell = ws3.cell(row=1, column=col, value=h)
        cell.font = hdr_font; cell.fill = hdr_fill
        cell.alignment = center; cell.border = thin
    events = Event.query.order_by(Event.created_at.desc()).limit(2000).all()
    for row, ev in enumerate(events, start=2):
        vals = [
            ev.session_id[:8],
            ev.event_type,
            ev.payload[:100] if ev.payload else "",
            ev.created_at.strftime("%d/%m/%Y %H:%M"),
        ]
        for col, val in enumerate(vals, 1):
            c = ws3.cell(row=row, column=col, value=val)
            c.font = Font(name="Arial", size=10); c.border = thin
    for col, w in enumerate([16,20,40,20], start=1):
        ws3.column_dimensions[get_column_letter(col)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"vert4_analytics_{datetime.utcnow().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(buf, as_attachment=True, download_name=filename,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

if __name__ == "__main__":
    print("\n✅ Servidor Flask VERT4® com Analytics — http://localhost:5000\n")
    app.run(host="0.0.0.0", port=5000)